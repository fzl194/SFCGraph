"""跨命令参数引用（references）+ 对象间引用（refers_to）构建。

读内网参数引用规则（csv/xlsx，GBK 兼容）→ 解析源/目标参数引用 →
  - parameter_references.jsonl：CommandParameter references 边（schema §4.4.2）
  - object_refers_to.jsonl：ConfigObject refers_to 边（从参数引用聚合推导，schema §4.5）

参数引用格式：`{nf}:{VERB_OBJECT}.{PARAM}`，如 `UDG:ADD_FLTBINDFLOWF.FILTERNAME`
（命令部分下划线连接 → 还原空格为 command_name；verb_OBJECT 第一个下划线分动词与对象）。
"""
import csv
from pathlib import Path

from .config_object import write_jsonl  # noqa: F401  （step 复用）

# 列名关键词（表头含说明也能匹配）
COL_KEYWORDS = {
    "nf": ("网元",),
    "version": ("版本",),
    "src_ref": ("配置", "后配"),      # 源参数引用 UDG:ADD_FLTBINDFLOWF.FILTERNAME
    "dst_ref": ("配置", "依赖参数"),  # 目标参数引用 UDG:ADD_FILTER.FILTERNAME
    "src_cond": ("是否匹配",),        # 源参数触发条件
    "check": ("索引检查",),           # 核查逻辑
    "binding": ("强绑定",),           # 0弱/1强（列名含"0：弱绑定，1：强绑定"）
    "cascade": ("是否联动删除",),     # 精确匹配，避开强弱绑定列说明里的"set需要联动删除"
}


def _load_rows(path):
    """读 csv（utf-8-sig 失败则 gb18030）或 xlsx（第一 sheet）→ (rows, header)。"""
    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_xlsx(p)
    for enc in ("utf-8-sig", "gb18030", "gbk"):
        try:
            with p.open(encoding=enc, newline="") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if not header:
                    return [], None
                rows = [r for r in reader if any((c or "").strip() for c in r)]
            return rows, header
        except UnicodeDecodeError:
            continue
    return [], None


def _load_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("读取 xlsx 需要安装 openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], None
    header = [(str(c) or "").strip() if c is not None else "" for c in rows[0]]
    data = [["" if v is None else str(v) for v in values]
            for values in rows[1:]
            if any(v not in (None, "") for v in values)]
    return data, header


def _find_col(header, keywords):
    for i, h in enumerate(header):
        if h and all(k in h for k in keywords):
            return i
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def _parse_param_ref(ref, version):
    """UDG:ADD_FLTBINDFLOWF.FILTERNAME → dict(parameter_id/object_name/...)，失败 None。"""
    if not ref or ":" not in ref:
        return None
    ref = ref.strip()
    ref_nf, _, rest = ref.partition(":")
    ref_nf = ref_nf.strip()
    if "." not in rest:
        return None
    cmd_part, _, param = rest.rpartition(".")
    cmd_part = cmd_part.strip()
    param = param.strip()
    if not cmd_part or not param or not ref_nf:
        return None
    verb, _, obj = cmd_part.partition("_")  # 第一个下划线分动词与对象
    command_name = f"{verb} {obj}" if obj else verb
    object_name = obj or verb
    return {
        "nf": ref_nf,
        "version": version,
        "command_name": command_name,
        "object_name": object_name,
        "param_name": param,
        "parameter_id": f"{ref_nf}@{version}@CommandParameter@{command_name}:{param}",
    }


def _strength(v):
    v = (v or "").strip()
    if v == "1":
        return "强绑定"
    if v == "0":
        return "弱绑定"
    return v or None


def build_param_references(path, default_version):
    """读参数引用规则 → {"parameter_references": [...], "object_refers_to": [...]}。"""
    rows, header = _load_rows(path)
    if not header:
        return {"parameter_references": [], "object_refers_to": []}
    cols = {k: _find_col(header, kw) for k, kw in COL_KEYWORDS.items()}

    refs = []
    obj_refs = {}  # (nf, ver, src_obj, dst_obj) -> set(via_param)
    for row in rows:
        ver = _cell(row, cols.get("version")) or default_version
        src = _parse_param_ref(_cell(row, cols.get("src_ref")), ver)
        dst = _parse_param_ref(_cell(row, cols.get("dst_ref")), ver)
        if not src or not dst:
            continue
        refs.append({
            "edge_type": "references",
            "from_parameter_ref": src["parameter_id"],
            "to_parameter_ref": dst["parameter_id"],
            "source_condition": _cell(row, cols.get("src_cond")) or None,
            "check_expression": _cell(row, cols.get("check")) or None,
            "binding_strength": _strength(_cell(row, cols.get("binding"))),
            "cascade_delete": _cell(row, cols.get("cascade")) == "1",
        })
        # 推导对象间 refers_to（同对象跳过；按 src_obj→dst_obj 去重，via_param 累积）
        if (src["object_name"] and dst["object_name"]
                and src["object_name"] != dst["object_name"]
                and src["nf"] == dst["nf"]):
            key = (src["nf"], ver, src["object_name"], dst["object_name"])
            obj_refs.setdefault(key, set()).add(src["param_name"])

    refers_to = [{
        "edge_type": "refers_to",
        "from_object_ref": f"{nf}@{ver}@ConfigObject@{sobj}",
        "to_object_ref": f"{nf}@{ver}@ConfigObject@{dobj}",
        "via_parameter": sorted(params),
    } for (nf, ver, sobj, dobj), params in obj_refs.items()]
    return {"parameter_references": refs, "object_refers_to": refers_to}
