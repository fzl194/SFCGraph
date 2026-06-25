"""CommandParameter 构建纯函数（CSV/XLSX → 参数节点 + 两类关系边）。

从旧 command-graph/builder/build_commandparameter.py 整体移植，逻辑不变：
  - 参数节点字段映射
  - 命令内条件依赖（来自"条件"列）
  - 可选的命令关联校验（对照 mml_commands.jsonl）
  - 占位行过滤（PLACEHOLDER_PARAMETER_NAMES / PLACEHOLDER_ID）

拆为独立模块以便 steps/parameter.py 调用、tests 直接导入，去掉 argparse main。
"""
import csv
import json
from pathlib import Path

from .constants import PLACEHOLDER_ID, PLACEHOLDER_PARAMETER_NAMES


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.lower() == "nan":
        return ""
    return text


def normalize_nullable_text(value):
    text = clean_text(value)
    if text in {"无", "null", "NULL", "None"}:
        return None
    return text or None


def parse_int(value):
    text = clean_text(value)
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def parse_bool(value):
    text = clean_text(value)
    if not text:
        return None
    if text in {"是", "true", "True", "1"}:
        return True
    if text in {"否", "false", "False", "0"}:
        return False
    return None


def parse_json_or_list(value):
    text = clean_text(value)
    if not text:
        return []
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        data = None
    if isinstance(data, list):
        return [clean_text(item) for item in data if clean_text(item)]
    parts = []
    for item in text.replace("\r", "\n").splitlines():
        for chunk in item.replace("，", ",").replace("；", ";").split(","):
            chunk = chunk.strip().strip('"')
            if chunk:
                parts.append(chunk)
    return parts


def parse_forbidden_values(value):
    items = parse_json_or_list(value)
    return items or None


def make_command_id(nf, version, command_name):
    return f"{nf}@{version}@MMLCommand@{command_name}"


def make_parameter_id(nf, version, command_name, parameter_name):
    return f"{nf}@{version}@CommandParameter@{command_name}:{parameter_name}"


def is_placeholder_row(row):
    parameter_name = clean_text(row.get("参数标识"))
    native_id = clean_text(row.get("参数ID"))
    return parameter_name.lower() in PLACEHOLDER_PARAMETER_NAMES or native_id == str(PLACEHOLDER_ID)


def parse_condition_map(text):
    raw = clean_text(text)
    if not raw:
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []

    items = []
    for key, required_mode in data.items():
        left, sep, right = str(key).partition("=")
        if not sep:
            continue
        source_native_id = parse_int(left)
        if source_native_id is None:
            continue
        items.append(
            {
                "source_native_id": source_native_id,
                "condition_value": clean_text(right),
                "required_mode": clean_text(required_mode),
            }
        )
    return items


def build_parameter(row, row_num, source_name):
    nf = clean_text(row.get("网元类型"))
    version = clean_text(row.get("网元版本"))
    command_name = clean_text(row.get("命令"))
    parameter_name = clean_text(row.get("参数标识"))
    native_parameter_id = parse_int(row.get("参数ID"))

    parameter = {
        "parameter_id": make_parameter_id(nf, version, command_name, parameter_name),
        "nf": nf,
        "version": version,
        "native_parameter_id": native_parameter_id,
        "command_id": make_command_id(nf, version, command_name),
        "parameter_name": parameter_name,
        "parameter_name_zh": normalize_nullable_text(row.get("参数名称")),
        "data_type": normalize_nullable_text(row.get("参数类型")),
        "required_mode": normalize_nullable_text(row.get("可必选")),
        "condition_for_required": normalize_nullable_text(row.get("条件")),
        "enum_values": parse_json_or_list(row.get("枚举")),
        "default_value": normalize_nullable_text(row.get("默认值")),
        "max_value": normalize_nullable_text(row.get("最大值")),
        "min_value": normalize_nullable_text(row.get("最小值")),
        "range_interval": normalize_nullable_text(row.get("区间")),
        "max_length": parse_int(row.get("最大长度")),
        "min_length": parse_int(row.get("最小长度")),
        "length": parse_int(row.get("长度")),
        "value_range": normalize_nullable_text(row.get("取值范围")),
        "bitfield": normalize_nullable_text(row.get("位域")),
        "string_format": normalize_nullable_text(row.get("字符串格式")),
        "regex_id": normalize_nullable_text(row.get("正则id")),
        "case_sensitive": parse_bool(row.get("是否区分大小写")),
        "comparison_operator": normalize_nullable_text(row.get("比较关系")),
        "forbidden_values": parse_forbidden_values(row.get("禁输值")),
        "conditional_range": normalize_nullable_text(row.get("条件区间")),
        "inheritance": normalize_nullable_text(row.get("继承关系")),
        "description": normalize_nullable_text(row.get("说明")),
    }

    return parameter


def build_has_parameter_edge(parameter):
    return {
        "edge_type": "has_parameter",
        "from_command_ref": parameter["command_id"],
        "to_parameter_ref": parameter["parameter_id"],
    }


def build_depends_on_edges(parameter_rows, source_name):
    edges = []
    unresolved = []
    by_command_and_native_id = {}

    for entry in parameter_rows:
        native_id = entry["parameter"].get("native_parameter_id")
        if native_id is None:
            continue
        by_command_and_native_id[(entry["parameter"]["command_id"], native_id)] = entry["parameter"]

    for entry in parameter_rows:
        parameter = entry["parameter"]
        conditions = parse_condition_map(entry["row"].get("条件"))
        for condition in conditions:
            source_key = (parameter["command_id"], condition["source_native_id"])
            source_parameter = by_command_and_native_id.get(source_key)
            if not source_parameter:
                unresolved.append(
                    {
                        "parameter_ref": parameter["parameter_id"],
                        "missing_source_native_id": condition["source_native_id"],
                        "condition_value": condition["condition_value"],
                    }
                )
                continue

            edges.append(
                {
                    "edge_type": "depends_on",
                    "from_parameter_ref": source_parameter["parameter_id"],
                    "to_parameter_ref": parameter["parameter_id"],
                    "condition_ref": source_parameter["parameter_name"],
                    "condition_logic": "等于",
                    "condition_value": condition["condition_value"],
                }
            )

    return edges, unresolved


def build_from_rows(rows, source_name, command_ids=None):
    parameters = []
    has_parameter_edges = []
    parameter_rows = []
    skipped_placeholders = 0
    missing_commands = []

    for index, row in enumerate(rows, start=2):
        if is_placeholder_row(row):
            skipped_placeholders += 1
            continue

        parameter = build_parameter(row, index, source_name)
        if command_ids is not None and parameter["command_id"] not in command_ids:
            missing_commands.append(parameter["command_id"])

        parameters.append(parameter)
        has_parameter_edges.append(build_has_parameter_edge(parameter))
        parameter_rows.append({"row_num": index, "row": row, "parameter": parameter})

    depends_on_edges, unresolved_dependencies = build_depends_on_edges(parameter_rows, source_name)

    return {
        "parameters": parameters,
        "has_parameter_edges": has_parameter_edges,
        "depends_on_edges": depends_on_edges,
        "skipped_placeholders": skipped_placeholders,
        "missing_commands": sorted(set(missing_commands)),
        "unresolved_dependencies": unresolved_dependencies,
    }


def load_csv_rows(path):
    with Path(path).open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def load_xlsx_rows(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 xlsx 需要安装 openpyxl。") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean_text(item) for item in rows[0]]
    records = []
    for values in rows[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else None
        if any(clean_text(value) for value in row.values()):
            records.append(row)
    return records


def load_rows(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return load_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx_rows(path)
    raise ValueError(f"暂不支持的输入格式: {suffix}")


def load_command_ids(path):
    command_ids = set()
    with Path(path).open(encoding="utf-8") as fh:
        for line in fh:
            raw = clean_text(line)
            if not raw:
                continue
            obj = json.loads(raw)
            command_id = clean_text(obj.get("command_id"))
            if command_id:
                command_ids.add(command_id)
    return command_ids


def write_jsonl(path, items):
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as fh:
        for item in items:
            fh.write(json.dumps(item, ensure_ascii=False) + "\n")
