"""core/seed.py 测试（用 openpyxl 造小 xlsx）。"""
import openpyxl

from builder.core.seed import extract_seed, UDG_NF_COLUMNS


def _make_xlsx(tmp_path):
    # 真实 xlsx 布局：col0=编号, col1=名称, col2=GGSN列(目录标记也写在该单元格), col3+=其余NF
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["特性编号", "特性名称", "GGSN(2G&3G)", "S/PGW-U(4G)", "S/PGW-U(5G NSA)", "UPF(5G)", "NB-IoT"])
    ws.append(["NFV基本特性", "", "", "", "", "", ""])   # section 行（仅首列有值）
    ws.append(["GWFD-000101", "云化分布式软件架构", "目录", "", "", "", ""])   # 目录行：col2=目录标记
    ws.append(["GWFD-000102", "微服务架构", "M", "M", "M", "M", "M"])          # 叶子：col2=GGSN=M
    p = tmp_path / "m.xlsx"
    wb.save(p)
    return p


def test_extract_seed_parses_leaf_and_dir(tmp_path):
    feats = extract_seed(_make_xlsx(tmp_path), sheets=[0], nf_columns=UDG_NF_COLUMNS)
    assert len(feats) == 2
    f1, f2 = feats
    assert f1["feature_code"] == "GWFD-000101" and f1["is_directory"] is True
    assert f2["feature_code"] == "GWFD-000102" and f2["is_directory"] is False
    assert f2["parent_feature_code"] == "GWFD-000101"  # 目录父追踪
    assert f2["catalog_section"] == "NFV基本特性"
    assert "GGSN(2G&3G)=M" in f2["nf_support_map"]
    assert "UPF(5G)=M" in f2["nf_support_map"]


def test_extract_seed_dedup(tmp_path):
    # 两 sheet 含同一特性 → 去重
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.append(["编号", "名称"])
    ws1.append(["GWFD-000101", "A", ""])
    ws2 = wb.create_sheet()
    ws2.append(["编号", "名称"])
    ws2.append(["GWFD-000101", "A", ""])
    p = tmp_path / "d.xlsx"
    wb.save(p)
    feats = extract_seed(p, sheets=[0, 1], nf_columns={})
    assert len(feats) == 1
