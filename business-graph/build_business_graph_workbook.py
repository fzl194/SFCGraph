from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
SOURCE_MD = BASE_DIR / "06-business-awareness-business-graph-final.md"
OUTPUT_XLSX = BASE_DIR / "06-business-awareness-business-graph-final.xlsx"


def read_lines(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8").splitlines()


def normalize_cell(text: str) -> str:
    value = text.strip()
    if value.startswith("`") and value.endswith("`"):
        value = value[1:-1]
    value = value.replace("<br>", "\n")
    value = value.replace("`", "")
    return value.strip()


def is_table_line(line: str) -> bool:
    stripped = line.strip()
    return stripped.startswith("|") and stripped.endswith("|")


def is_separator_row(cells: list[str]) -> bool:
    return all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells)


def parse_markdown_table(table_lines: list[str]) -> list[list[str]]:
    rows: list[list[str]] = []
    for line in table_lines:
        parts = [normalize_cell(part) for part in line.strip()[1:-1].split("|")]
        rows.append(parts)
    if len(rows) >= 2 and is_separator_row(rows[1]):
        rows.pop(1)
    return rows


def find_line_index(lines: list[str], needle: str) -> int:
    for i, line in enumerate(lines):
        if needle in line:
            return i
    raise ValueError(f"Cannot find section containing: {needle}")


def next_heading_index(lines: list[str], start_idx: int) -> int:
    for i in range(start_idx + 1, len(lines)):
        if lines[i].startswith("## ") or lines[i].startswith("### "):
            return i
    return len(lines)


def extract_first_table_after(lines: list[str], start_idx: int, end_idx: int | None = None) -> list[list[str]]:
    if end_idx is None:
        end_idx = len(lines)
    i = start_idx
    while i < end_idx:
        if is_table_line(lines[i]):
            block: list[str] = []
            while i < end_idx and is_table_line(lines[i]):
                block.append(lines[i])
                i += 1
            return parse_markdown_table(block)
        i += 1
    raise ValueError(f"No table found after line {start_idx}")


def transpose_field_matrix(table: list[list[str]]) -> list[dict[str, str]]:
    headers = [normalize_cell(cell) for cell in table[0]]
    instances = headers[1:]
    result = [{headers[0]: name} for name in instances]
    for row in table[1:]:
        field = normalize_cell(row[0])
        for idx, value in enumerate(row[1:]):
            result[idx][field] = value
    return result


def kv_table_to_dict(table: list[list[str]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in table[1:]:
        if len(row) >= 2:
            result[normalize_cell(row[0])] = row[1]
    return result


def write_sheet(ws, headers: list[str], rows: Iterable[dict[str, object] | list[object]]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(header, "") for header in headers])
        else:
            ws.append(list(row))

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def build_object_sheet_rows(lines: list[str]) -> list[dict[str, str]]:
    schema_table = extract_first_table_after(lines, find_line_index(lines, "### 4.1 schema 总表"))
    section_map = {
        "BusinessDomain": "7",
        "NetworkScenario": "8",
        "DeliverySolution": "10",
        "EngineeringTask": "11",
        "Participant": "12",
        "Scope": "13",
        "DecisionPoint": "14",
        "ValidationPlan": "16",
        "BusinessRule": "17",
        "DomainSemanticObject": "20",
        "Feature": "21",
        "ConfigObject": "22",
    }
    rows: list[dict[str, str]] = []
    for row in schema_table[1:]:
        if len(row) < 6:
            continue
        obj = normalize_cell(row[0])
        rows.append(
            {
                "schema对象": obj,
                "中文对象名": row[1],
                "所属层": row[2],
                "定义": row[3],
                "在图谱中的作用": row[4],
                "需要重点展开的业务含义": row[5],
                "对应章节": section_map.get(obj, ""),
            }
        )
    return rows


def build_edge_sheet_rows(lines: list[str]) -> list[dict[str, str]]:
    table = extract_first_table_after(lines, find_line_index(lines, "### 5.1 schema 关系总表"))
    rows: list[dict[str, str]] = []
    for row in table[1:]:
        if len(row) < 5:
            continue
        rows.append(
            {
                "起点schema": row[0],
                "终点schema": row[1],
                "关系标识": row[2],
                "中文关系名": row[3],
                "业务含义": row[4],
            }
        )
    return rows


def build_business_domain_rows(lines: list[str]) -> list[dict[str, str]]:
    table = extract_first_table_after(lines, find_line_index(lines, "### 7.2"))
    row = kv_table_to_dict(table)
    return [row]


def build_transposed_instance_rows(lines: list[str], heading_token: str) -> list[dict[str, str]]:
    table = extract_first_table_after(lines, find_line_index(lines, heading_token))
    return transpose_field_matrix(table)


def build_standard_table_rows(lines: list[str], heading_token: str) -> list[dict[str, str]]:
    table = extract_first_table_after(lines, find_line_index(lines, heading_token))
    headers = table[0]
    return [dict(zip(headers, row)) for row in table[1:]]


def build_grouped_table_rows(lines: list[str], groups: list[tuple[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for group_name, token in groups:
        table = extract_first_table_after(lines, find_line_index(lines, token))
        headers = table[0]
        for row in table[1:]:
            item = {"分组": group_name}
            for idx, header in enumerate(headers):
                item[header] = row[idx] if idx < len(row) else ""
            rows.append(item)
    return rows


def build_task_sequences(lines: list[str]) -> list[dict[str, str]]:
    start = find_line_index(lines, "#### 各方案任务编排顺序")
    end = find_line_index(lines, "### 11.3 原子 Task 详情")
    section = lines[start:end]
    rows: list[dict[str, str]] = []
    current_ds = None
    buffer: list[str] = []
    for line in section:
        stripped = line.strip()
        if stripped.startswith("**DS-"):
            if current_ds and buffer:
                rows.append({"方案": current_ds, "任务编排顺序": " ".join(buffer).strip()})
                buffer = []
            current_ds = stripped.strip("*")
            current_ds = current_ds[:-1] if current_ds.endswith("：") else current_ds
        elif stripped.startswith("T-") or stripped.startswith("→ T-"):
            buffer.append(stripped)
    if current_ds and buffer:
        rows.append({"方案": current_ds, "任务编排顺序": " ".join(buffer).strip()})
    return rows


def build_task_detail_rows(lines: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    task_heading_pattern = re.compile(r"^#### (T-[A-Z]+-\d{3})\s+(.+)$")
    indices = []
    for i, line in enumerate(lines):
        match = task_heading_pattern.match(line.strip())
        if match:
            indices.append((i, match.group(1), match.group(2)))
    for idx, task_id, task_name in indices:
        next_idx = next((i for i, _, _ in indices if i > idx), len(lines))
        table = extract_first_table_after(lines, idx, next_idx)
        item = kv_table_to_dict(table)
        if "task_id" not in item:
            item["task_id"] = task_id
        if "task_name" not in item:
            item["task_name"] = task_name
        rows.append(item)
    return rows


def build_instance_relation_rows(lines: list[str]) -> list[dict[str, str]]:
    sections = [
        ("23.1 业务层关系", "### 23.1"),
        ("23.2 方案到任务关系", "### 23.2"),
        ("23.3 方案到支撑对象关系", "### 23.3"),
        ("23.4 支撑层内部关系", "### 23.4"),
    ]
    rows: list[dict[str, str]] = []
    for group_name, token in sections:
        table = extract_first_table_after(lines, find_line_index(lines, token))
        headers = table[0]
        for row in table[1:]:
            item = {"关系分组": group_name}
            for idx, header in enumerate(headers):
                item[header] = row[idx] if idx < len(row) else ""
            rows.append(item)
    return rows


def main() -> None:
    lines = read_lines(SOURCE_MD)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets: list[tuple[str, list[str], list[dict[str, object]]]] = [
        (
            "子段管理",
            ["schema对象", "中文对象名", "所属层", "定义", "在图谱中的作用", "需要重点展开的业务含义", "对应章节"],
            build_object_sheet_rows(lines),
        ),
        (
            "边管理",
            ["起点schema", "终点schema", "关系标识", "中文关系名", "业务含义"],
            build_edge_sheet_rows(lines),
        ),
        (
            "BusinessDomain",
            ["domain_id", "domain_name", "domain_summary", "status"],
            build_business_domain_rows(lines),
        ),
        (
            "NetworkScenario",
            ["字段", "scenario_id", "scenario_name", "scenario_summary", "judgment_basis", "typical_outcome", "status"],
            build_transposed_instance_rows(lines, "### 8.2"),
        ),
        (
            "DeliverySolution",
            ["字段", "solution_id", "solution_name", "solution_summary", "core_mechanism_combo", "source_evidence_ids", "status"],
            build_transposed_instance_rows(lines, "### 10.2"),
        ),
        (
            "Task编排",
            ["方案", "任务编排顺序"],
            build_task_sequences(lines),
        ),
        (
            "Task详情",
            [
                "task_id",
                "task_name",
                "task_summary",
                "phase",
                "input_artifacts",
                "output_artifacts",
                "command",
                "config_object",
                "source_evidence_ids",
                "status",
            ],
            build_task_detail_rows(lines),
        ),
        (
            "Participant",
            ["participant_id", "participant_name", "participant_type", "plane_or_side", "responsibility_summary", "status"],
            build_standard_table_rows(lines, "### 12.2"),
        ),
        (
            "Scope",
            ["scope_id", "scope_name", "scope_type", "scope_expression", "scope_summary", "status"],
            build_standard_table_rows(lines, "### 13.2"),
        ),
        (
            "DecisionPoint",
            ["decision_id", "decision_name", "挂载层级", "decision_question", "option_set", "trigger_condition", "status"],
            build_standard_table_rows(lines, "### 14.2"),
        ),
        (
            "ValidationPlan",
            ["validation_id", "validation_name", "对应方案", "validation_goal", "target_objects", "pass_condition", "expected_result", "status"],
            build_standard_table_rows(lines, "### 16.2"),
        ),
        (
            "BusinessRule",
            ["rule_id", "rule_name", "rule_type", "对应方案/计划", "check_target", "check_logic", "expected_result / violation_effect / next_action", "status"],
            build_standard_table_rows(lines, "### 17.2"),
        ),
        (
            "DomainSemanticObject",
            ["语义对象", "定义", "在业务感知中的业务含义", "典型落点"],
            build_standard_table_rows(lines, "### 20.1"),
        ),
        (
            "Feature",
            ["特性", "所属侧别", "正式能力定位", "在业务图谱中的作用"],
            build_standard_table_rows(lines, "### 21.1"),
        ),
        (
            "ConfigObject",
            ["对象链位置", "关键对象", "所属侧别", "业务作用"],
            build_standard_table_rows(lines, "### 22.1"),
        ),
        (
            "关系实例",
            ["关系分组", "关系标识", "中文关系名", "起点对象", "终点对象", "业务解释"],
            build_instance_relation_rows(lines),
        ),
    ]

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title)
        write_sheet(ws, headers, rows)

    wb.save(OUTPUT_XLSX)
    print(str(OUTPUT_XLSX))


if __name__ == "__main__":
    main()
